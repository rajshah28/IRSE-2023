from openpyxl import Workbook

code_snippets = []
comments = []
with open('sample.c', 'r') as file:
    lines = file.readlines()
    l = len(lines)
    i = 0
    i_list = [0]*l
    for i in range(l):
        if i_list[i]==1:
            continue
        words = lines[i].strip().split()
        if words:
            if lines[i][:2]=='//':
                comment = ''
                comment+=lines[i]
                word = lines[i+1].strip().split()
                t = i
                if word:
                    if lines[i+1][:2]=='//':
                        for k in range(t+1,l):
                            i_list[k]=1
                            w = lines[k].strip().split()
                            if w:
                                if lines[k][:2]=='//':
                                    comment+=lines[k]
                                    i+=1
                                    # print("1")
                                else:
                                    break
                code = ''
                for j in range(i+1, l):
                    word = lines[j].strip().split()
                    if word:
                        if lines[j][:2]=='//':
                            break
                        code+=lines[j]
                code_snippets.append(code)
                comments.append(comment)
        i+=1

# print(comments)
# print(code_snippets)
workbook = Workbook()
sheet = workbook.active
sheet.column_dimensions["A"].width = 50
sheet.column_dimensions["B"].width = 50

for i, (code_snippet, comment) in enumerate(zip(code_snippets, comments)):
    sheet.cell(row=i + 1, column=1, value=code_snippet)
    sheet.cell(row=i + 1, column=2, value=comment)

workbook.save('scrape.xlsx')

